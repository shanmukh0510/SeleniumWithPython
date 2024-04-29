# Dta driven testing

# 1) read data from exel
# 2) how to write data into exel
# 3) data driven test case

# reading
# data = sheet.cell(r,c).value

# writing
# data = sheet.cell(r,c).value = "welcome"


import openpyxl
#file -->workbook-->sheets--->rows--->cells
#   read exel data
file = "D:\Financial Sample.xlsx"

workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = sheet.max_row
columns = sheet.max_column

print(rows)
print(columns)
for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value,end= "       ")
    print()


